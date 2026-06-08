#!/usr/bin/env python3
"""
Generate Master Sheets for 4-Psychiatrist Rating Study
=======================================================

Creates:
  1. Progress Tracker — track submission status per psychiatrist
  2. Case Metadata Reference — all 100 cases with anonymized info
  3. Rating Consolidation Template — combine all 4 ratings post-submission
  4. Analysis Sheet — full analysis with agreement metrics

All placed in: 04_results/human_validation/master_sheets/
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
SOURCE_CSV = BASE / "04_results/human_validation/rater_sample_400_blinded.csv"
KEY_CSV = BASE / "04_results/human_validation/rater_sample_400_key.csv"
OUTPUT_DIR = BASE / "04_results/human_validation/master_sheets"

# ── Styles ──────────────────────────────────────────────────────────────────
TITLE_FONT = Font(bold=True, size=16, color="2F5496")
SECTION_FONT = Font(bold=True, size=13, color="2F5496")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
BODY_FONT = Font(size=11)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ALT_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")


def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=THIN):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    else: cell.alignment = WRAP_ALIGN
    cell.border = border
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Progress Tracker
# ═══════════════════════════════════════════════════════════════════════════════

def create_progress_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress Tracker"

    # Title
    ws.merge_cells("A1:F1")
    set_cell(ws, 1, 1, "4-PSYCHIATRIST RATING STUDY — PROGRESS TRACKER",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    ws.merge_cells("A2:F2")
    set_cell(ws, 2, 1, "PAHS LLM Hallucination Study 2026",
             font=Font(size=12, italic=True, color="2F5496"), align=CENTER_ALIGN)

    # Study info
    info_row = 4
    set_cell(ws, info_row, 1, "Study Design:", font=Font(bold=True, size=11))
    set_cell(ws, info_row, 2, "100 cases × 4 psychiatrists = 400 total ratings")
    set_cell(ws, info_row+1, 1, "Cases per psychiatrist:", font=Font(bold=True, size=11))
    set_cell(ws, info_row+1, 2, "100 (~1–2 hours each)")
    set_cell(ws, info_row+2, 1, "Models:", font=Font(bold=True, size=11))
    set_cell(ws, info_row+2, 2, "4 anonymized models (A, B, C, D)")
    set_cell(ws, info_row+3, 1, "Stratification:", font=Font(bold=True, size=11))
    set_cell(ws, info_row+3, 2, "4 models × 3 conditions × 2 lengths = 24 cells (4–5 per cell)")

    # Psychiatrist tracking table
    header_row = 10
    set_cell(ws, header_row, 1, "PSYCHIATRIST", font=Font(bold=True, size=12, color="2F5496"))
    ws.merge_cells(f"A{header_row}:F{header_row}")

    headers = ["Psychiatrist", "Excel File", "Status", "Date Started", "Date Submitted", "Notes"]
    for ci, h in enumerate(headers, 1):
        set_cell(ws, header_row+1, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    statuses = ["Not started", "In progress", "Submitted", "Validated"]
    psychiatrists = [
        ("Psychiatrist 1", "Psychiatrist_1_rating_sheet.xlsx"),
        ("Psychiatrist 2", "Psychiatrist_2_rating_sheet.xlsx"),
        ("Psychiatrist 3", "Psychiatrist_3_rating_sheet.xlsx"),
        ("Psychiatrist 4", "Psychiatrist_4_rating_sheet.xlsx"),
    ]

    for i, (name, file) in enumerate(psychiatrists):
        r = header_row + 2 + i
        fill = [GREEN_FILL, YELLOW_FILL, BLUE_FILL, GRAY_FILL][i % 4] if i == 0 else ALT_FILL if i % 2 == 0 else None
        set_cell(ws, r, 1, name, font=Font(bold=True, size=11), fill=fill)
        set_cell(ws, r, 2, file, fill=fill)
        set_cell(ws, r, 3, "Not started", fill=YELLOW_FILL if fill else ALT_FILL)
        set_cell(ws, r, 4, "", fill=fill)
        set_cell(ws, r, 5, "", fill=fill)
        set_cell(ws, r, 6, "", fill=fill)

    # Summary section
    sum_row = header_row + 8
    set_cell(ws, sum_row, 1, "SUMMARY", font=Font(bold=True, size=12, color="2F5496"))
    ws.merge_cells(f"A{sum_row}:F{sum_row}")

    summary_items = [
        ("Total cases", "100"),
        ("Total ratings expected", "400"),
        ("Ratings received", "=COUNTIF(C12:C15,\"Submitted\")*100"),
        ("Completion %", "=COUNTIF(C12:C15,\"Submitted\")/4"),
        ("", ""),
        ("Next steps after all submitted:", ""),
        ("1. Export each Excel as CSV", ""),
        ("2. Run: python calculate_kappa_4raters.py *.csv", ""),
        ("3. Use Rating Consolidation Template to combine", ""),
    ]

    for i, (label, val) in enumerate(summary_items):
        r = sum_row + 1 + i
        set_cell(ws, r, 1, label, font=Font(bold=True, size=11) if label else BODY_FONT)
        set_cell(ws, r, 2, val)

    # Column widths
    widths = [20, 40, 18, 16, 16, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A12"

    out = OUTPUT_DIR / "01_progress_tracker.xlsx"
    wb.save(out)
    print(f"  ✅ 01_progress_tracker.xlsx")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Case Metadata Reference
# ═══════════════════════════════════════════════════════════════════════════════

def create_case_metadata_reference():
    df = pd.read_csv(SOURCE_CSV)
    df_key = pd.read_csv(KEY_CSV)

    wb = Workbook()
    ws = wb.active
    ws.title = "Case Metadata"

    # Title
    ws.merge_cells("A1:H1")
    set_cell(ws, 1, 1, "CASE METADATA REFERENCE — 100 Cases",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    ws.merge_cells("A2:H2")
    set_cell(ws, 2, 1, "For reference during rating. Model names are anonymized (A/B/C/D).",
             font=Font(size=11, italic=True), align=CENTER_ALIGN)

    # Headers
    headers = ["Case_ID", "Orig_Case_ID", "Model", "Token", "Condition", "Length", "Category", "Notes"]
    for ci, h in enumerate(headers, 1):
        set_cell(ws, 4, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), start=5):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        set_cell(ws, row_idx, 1, row["case_seq_id"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 2, row["case_id"], fill=fill)
        set_cell(ws, row_idx, 3, row["model"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 4, row.get("target_token", ""), fill=fill)
        set_cell(ws, row_idx, 5, row["condition"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 6, row["vignette_length"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 7, row.get("category", ""), fill=fill)
        set_cell(ws, row_idx, 8, "", fill=fill)  # Empty notes column

    # Column widths
    widths = [10, 25, 12, 35, 22, 10, 25, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{len(df) + 4}"

    # Model key (at bottom, clearly marked as SECRET)
    key_row = len(df) + 7
    ws.merge_cells(f"A{key_row}:H{key_row}")
    set_cell(ws, key_row, 1, "⚠️ MODEL ANONYMIZATION KEY — KEEP SECRET UNTIL ALL RATINGS COMPLETE ⚠️",
             font=Font(bold=True, size=12, color="C00000"), fill=RED_FILL, align=CENTER_ALIGN)

    for i, (_, row_k) in enumerate(df_key.iterrows()):
        r = key_row + 1 + i
        set_cell(ws, r, 1, row_k["anonymized_model"], font=Font(bold=True, size=11), fill=RED_FILL)
        set_cell(ws, r, 2, row_k["actual_model"], fill=RED_FILL)

    out = OUTPUT_DIR / "02_case_metadata_reference.xlsx"
    wb.save(out)
    print(f"  ✅ 02_case_metadata_reference.xlsx")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Rating Consolidation Template
# ═══════════════════════════════════════════════════════════════════════════════

def create_consolidation_template():
    df = pd.read_csv(SOURCE_CSV)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidation"

    # Title
    ws.merge_cells("A1:M1")
    set_cell(ws, 1, 1, "RATING CONSOLIDATION — Combine All 4 Psychiatrist Ratings",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    ws.merge_cells("A2:M2")
    set_cell(ws, 2, 1,
             "Paste each psychiatrist's Hallucination (0/1) and Confidence (1–5) ratings below, "
             "then use the Analysis Sheet for Kappa calculation.",
             font=Font(size=11, italic=True), align=CENTER_ALIGN)

    # Instructions
    set_cell(ws, 4, 1, "INSTRUCTIONS:", font=SECTION_FONT)
    ws.merge_cells("A4:M4")
    set_cell(ws, 5, 1,
             "1. Export each psychiatrist's Excel Rating Sheet as CSV. "
             "2. For each case (by Case_ID), enter the Hallucination rating (0/1) and Confidence (1–5) "
             "from each psychiatrist. 3. The Agreement column will auto-calculate. "
             "4. Then open the Analysis Sheet for full statistics.",
             font=BODY_FONT)

    # Headers
    headers = [
        "Case_ID", "Model", "Token", "Condition", "Length",
        "Psych_1_Rating", "Psych_1_Conf",
        "Psych_2_Rating", "Psych_2_Conf",
        "Psych_3_Rating", "Psych_3_Conf",
        "Psych_4_Rating", "Psych_4_Conf",
        "Agreement", "Notes"
    ]
    for ci, h in enumerate(headers, 1):
        fill = HEADER_FILL if ci <= 5 else PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        set_cell(ws, 7, ci, h, font=HEADER_FONT, fill=fill, align=CENTER_ALIGN)

    # Data rows — pre-fill case info, leave rating columns empty
    for row_idx, (_, row) in enumerate(df.iterrows(), start=8):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        set_cell(ws, row_idx, 1, row["case_seq_id"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 2, row["model"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 3, row.get("target_token", ""), fill=fill)
        set_cell(ws, row_idx, 4, row["condition"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws, row_idx, 5, row["vignette_length"], fill=fill, align=CENTER_ALIGN)

        # Empty rating columns (columns 5-12)
        # Empty rating columns (columns 6-13)
        for ci in range(6, 14):
            set_cell(ws, row_idx, ci, "", fill=YELLOW_FILL if ci % 2 == 0 else BLUE_FILL)

        # Agreement formula column (col 14)
        agreement_cell = ws.cell(row=row_idx, column=14)
        agreement_cell.value = f'=IF(OR(F{row_idx}="","G{row_idx}="","H{row_idx}="","I{row_idx}="","J{row_idx}="","K{row_idx}="","L{row_idx}="","M{row_idx}=""),"",IF(AND(F{row_idx}=G{row_idx},G{row_idx}=H{row_idx},H{row_idx}=I{row_idx},I{row_idx}=J{row_idx},J{row_idx}=K{row_idx},K{row_idx}=L{row_idx},L{row_idx}=M{row_idx}),"Unanimous",IF(OR(AND(F{row_idx}=G{row_idx},G{row_idx}=H{row_idx}),AND(G{row_idx}=H{row_idx},H{row_idx}=I{row_idx}),AND(F{row_idx}=H{row_idx},H{row_idx}=I{row_idx}),AND(F{row_idx}=G{row_idx},G{row_idx}=I{row_idx})),"Majority (3/4)","Split")))'
        agreement_cell.alignment = CENTER_ALIGN
        agreement_cell.font = Font(size=11, bold=True)
        agreement_cell.border = THIN

        set_cell(ws, row_idx, 15, "", fill=GRAY_FILL)

    # Column widths
    widths = [10, 12, 35, 22, 10, 14, 12, 14, 12, 14, 12, 14, 12, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:O{len(df) + 7}"

    out = OUTPUT_DIR / "03_rating_consolidation_template.xlsx"
    wb.save(out)
    print(f"  ✅ 03_rating_consolidation_template.xlsx")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Analysis Sheet
# ═══════════════════════════════════════════════════════════════════════════════

def create_analysis_sheet():
    df = pd.read_csv(SOURCE_CSV)
    df_key = pd.read_csv(KEY_CSV)

    wb = Workbook()

    # ── Tab 1: Overview ──────────────────────────────────────────────────────
    ws_overview = wb.active
    ws_overview.title = "Overview"

    ws_overview.merge_cells("A1:H1")
    set_cell(ws_overview, 1, 1, "ANALYSIS SHEET — 4-Psychiatrist Inter-Rater Reliability",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    # Study summary
    row = 3
    set_cell(ws_overview, row, 1, "STUDY SUMMARY", font=SECTION_FONT)
    ws_overview.merge_cells(f"A{row}:H{row}")

    summary_data = [
        ("Total unique cases", "100"),
        ("Psychiatrists", "4"),
        ("Total ratings", "400"),
        ("Cases per psychiatrist", "100"),
        ("Models (anonymized)", "A, B, C, D"),
        ("Conditions", "DEFAULT, DETERMINISTIC, SAFETY_INSTRUCTION"),
        ("Lengths", "short, long"),
        ("Cells (model×condition×length)", "24"),
        ("Cases per cell", "4–5"),
    ]
    for i, (label, val) in enumerate(summary_data):
        r = row + 1 + i
        set_cell(ws_overview, r, 1, label, font=Font(bold=True, size=11))
        set_cell(ws_overview, r, 2, val)

    # Analysis plan
    row = row + len(summary_data) + 2
    set_cell(ws_overview, row, 1, "ANALYSIS PLAN", font=SECTION_FONT)
    ws_overview.merge_cells(f"A{row}:H{row}")

    plan_data = [
        ("1. Pairwise Cohen's Kappa", "6 pairwise comparisons between all 4 psychiatrists"),
        ("2. Fleiss' Kappa", "Overall multi-rater agreement across all 4"),
        ("3. Per-case agreement", "Unanimous (4/4), Majority (3/4), Split (2/2)"),
        ("4. Prevalence", "% rated as hallucination by each psychiatrist"),
        ("5. Confidence analysis", "Mean confidence per psychiatrist and by agreement level"),
        ("6. Stratified analysis", "Kappa by model, condition, length (exploratory, small n)"),
    ]
    for i, (label, val) in enumerate(plan_data):
        r = row + 1 + i
        set_cell(ws_overview, r, 1, label, font=Font(bold=True, size=11))
        set_cell(ws_overview, r, 2, val)
        ws_overview.merge_cells(f"B{r}:H{r}")

    # Kappa interpretation guide
    row = row + len(plan_data) + 2
    set_cell(ws_overview, row, 1, "KAPPA INTERPRETATION GUIDE (Landis & Koch, 1977)", font=SECTION_FONT)
    ws_overview.merge_cells(f"A{row}:H{row}")

    kappa_guide = [
        ("< 0.00", "Poor", "RED"),
        ("0.00 – 0.20", "Slight", "ORANGE"),
        ("0.21 – 0.40", "Fair", "YELLOW"),
        ("0.41 – 0.60", "Moderate", "LIGHT GREEN"),
        ("0.61 – 0.80", "Substantial", "GREEN"),
        ("0.81 – 1.00", "Almost perfect", "DARK GREEN"),
    ]
    fills = [RED_FILL, YELLOW_FILL, YELLOW_FILL, GREEN_FILL, GREEN_FILL, GREEN_FILL]
    for i, (range_val, interp, _) in enumerate(kappa_guide):
        r = row + 1 + i
        set_cell(ws_overview, r, 1, range_val, font=Font(bold=True, size=11), fill=fills[i], align=CENTER_ALIGN)
        set_cell(ws_overview, r, 2, interp, fill=fills[i], align=CENTER_ALIGN)

    for i, w in enumerate([25, 20, 20, 20, 20, 20, 20, 20], 1):
        ws_overview.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 2: Pairwise Kappa Template ──────────────────────────────────────
    ws_pairwise = wb.create_sheet("Pairwise Kappa")

    ws_pairwise.merge_cells("A1:F1")
    set_cell(ws_pairwise, 1, 1, "PAIRWISE COHEN'S KAPPA — Template",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    headers = ["Rater Pair", "N", "κ", "95% CI Lower", "95% CI Upper", "Interpretation"]
    for ci, h in enumerate(headers, 1):
        set_cell(ws_pairwise, 3, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    pairs = [
        ("Psychiatrist 1 vs 2",), ("Psychiatrist 1 vs 3",), ("Psychiatrist 1 vs 4",),
        ("Psychiatrist 2 vs 3",), ("Psychiatrist 2 vs 4",), ("Psychiatrist 3 vs 4",),
    ]
    for i, (pair,) in enumerate(pairs):
        r = 4 + i
        fill = ALT_FILL if i % 2 == 0 else None
        set_cell(ws_pairwise, r, 1, pair, font=Font(bold=True, size=11), fill=fill)
        for ci in range(2, 7):
            set_cell(ws_pairwise, r, ci, "", fill=fill)

    # Summary row
    r = 11
    set_cell(ws_pairwise, r, 1, "Mean κ", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r, 2, "=AVERAGE(B4:B9)", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r+1, 1, "Median κ", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r+1, 2, "=MEDIAN(B4:B9)", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r+2, 1, "Min κ", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r+2, 2, "=MIN(B4:B9)", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r+3, 1, "Max κ", font=Font(bold=True, size=12))
    set_cell(ws_pairwise, r+3, 2, "=MAX(B4:B9)", font=Font(bold=True, size=12))

    for i, w in enumerate([30, 8, 12, 14, 14, 18], 1):
        ws_pairwise.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 3: Fleiss Kappa Template ─────────────────────────────────────────
    ws_fleiss = wb.create_sheet("Fleiss Kappa")

    ws_fleiss.merge_cells("A1:D1")
    set_cell(ws_fleiss, 1, 1, "FLEISS' KAPPA — Multi-Rater Agreement",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    fleiss_items = [
        ("Cases rated", "100"),
        ("Raters per case", "4"),
        ("Fleiss' κ", "—"),
        ("95% CI Lower", "—"),
        ("95% CI Upper", "—"),
        ("Observed agreement (P̄)", "—"),
        ("Expected agreement (Pe)", "—"),
        ("Interpretation", "—"),
        ("", ""),
        ("Category", "Proportion"),
        ("0 (No hallucination)", "—"),
        ("1 (Hallucination)", "—"),
    ]
    for i, (label, val) in enumerate(fleiss_items):
        r = 3 + i
        if label == "":
            continue
        set_cell(ws_fleiss, r, 1, label, font=Font(bold=True, size=11) if not label.startswith("(") else BODY_FONT)
        set_cell(ws_fleiss, r, 2, val)

    for i, w in enumerate([30, 20, 20, 20], 1):
        ws_fleiss.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 4: Per-Case Agreement ────────────────────────────────────────────
    ws_agreement = wb.create_sheet("Per-Case Agreement")

    ws_agreement.merge_cells("A1:G1")
    set_cell(ws_agreement, 1, 1, "PER-CASE AGREEMENT SUMMARY",
             font=TITLE_FONT, fill=BLUE_FILL, align=CENTER_ALIGN)

    headers = ["Case_ID", "Model", "Token", "Condition", "Length", "N Agree (out of 4)", "Agreement Level"]
    for ci, h in enumerate(headers, 1):
        set_cell(ws_agreement, 3, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        set_cell(ws_agreement, row_idx, 1, row["case_seq_id"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws_agreement, row_idx, 2, row["model"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws_agreement, row_idx, 3, row.get("target_token", ""), fill=fill)
        set_cell(ws_agreement, row_idx, 4, row["condition"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws_agreement, row_idx, 5, row["vignette_length"], fill=fill, align=CENTER_ALIGN)
        set_cell(ws_agreement, row_idx, 6, "", fill=YELLOW_FILL, align=CENTER_ALIGN)
        set_cell(ws_agreement, row_idx, 7, "", fill=GRAY_FILL, align=CENTER_ALIGN)

    for i, w in enumerate([10, 12, 35, 22, 10, 18, 20], 1):
        ws_agreement.column_dimensions[get_column_letter(i)].width = w

    ws_agreement.freeze_panes = "A4"
    ws_agreement.auto_filter.ref = f"A3:G{len(df) + 3}"

    # ── Tab 5: Model Key (Secret) ────────────────────────────────────────────
    ws_key = wb.create_sheet("Model Key (SECRET)")

    ws_key.merge_cells("A1:B1")
    set_cell(ws_key, 1, 1, "⚠️ MODEL ANONYMIZATION KEY — KEEP SECRET ⚠️",
             font=Font(bold=True, size=14, color="C00000"), fill=RED_FILL, align=CENTER_ALIGN)

    set_cell(ws_key, 3, 1, "Anonymous ID", font=HEADER_FONT, fill=RED_FILL, align=CENTER_ALIGN)
    set_cell(ws_key, 3, 2, "Real Model Name", font=HEADER_FONT, fill=RED_FILL, align=CENTER_ALIGN)

    for i, (_, row_k) in enumerate(df_key.iterrows()):
        r = 4 + i
        set_cell(ws_key, r, 1, row_k["anonymized_model"], font=Font(bold=True, size=12), fill=RED_FILL)
        set_cell(ws_key, r, 2, row_k["actual_model"], fill=RED_FILL)

    ws_key.column_dimensions["A"].width = 20
    ws_key.column_dimensions["B"].width = 50

    out = OUTPUT_DIR / "04_analysis_sheet.xlsx"
    wb.save(out)
    print(f"  ✅ 04_analysis_sheet.xlsx")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("GENERATING MASTER SHEETS")
    print("=" * 60)

    create_progress_tracker()
    create_case_metadata_reference()
    create_consolidation_template()
    create_analysis_sheet()

    print()
    print("=" * 60)
    print("ALL MASTER SHEETS GENERATED")
    print(f"Location: {OUTPUT_DIR}")
    print("=" * 60)
    print()
    print("Files:")
    print("  01_progress_tracker.xlsx          — Track submission status")
    print("  02_case_metadata_reference.xlsx   — Case info (give to psychiatrists)")
    print("  03_rating_consolidation_template.xlsx — Combine ratings post-submission")
    print("  04_analysis_sheet.xlsx            — Full analysis with Kappa templates")
    print()
    print("Workflow:")
    print("  1. Give each psychiatrist their rating sheet + case metadata reference")
    print("  2. Track progress with progress tracker")
    print("  3. After all submitted, paste ratings into consolidation template")
    print("  4. Use analysis sheet for Kappa and agreement statistics")
    print("  5. Run calculate_kappa_4raters.py for automated Kappa calculation")


if __name__ == "__main__":
    main()
