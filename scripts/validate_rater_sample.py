#!/usr/bin/env python3
"""Validate the 100-case rater sample and Excel files."""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

OUTPUT_DIR = Path("04_results/human_validation")

print("=" * 60)
print("VALIDATION: 100-Case Rater Sample")
print("=" * 60)

# 1. File existence
files = [
    "rater_sample_100.csv",
    "rater_sample_100_blinded.csv",
    "rater_sample_100_key.csv",
    "rater_sample_100_summary.txt",
    "Psychiatrist_1_rating_sheet.xlsx",
    "Psychiatrist_2_rating_sheet.xlsx",
    "Psychiatrist_3_rating_sheet.xlsx",
    "Psychiatrist_4_rating_sheet.xlsx",
]
print("\n[1] File check:")
for f in files:
    p = OUTPUT_DIR / f
    status = "OK" if p.exists() else "MISSING"
    size = p.stat().st_size if p.exists() else 0
    print(f"  {status}: {f} ({size:,} bytes)")

# 2. Blinded data
df = pd.read_csv(OUTPUT_DIR / "rater_sample_100_blinded.csv")
print(f"\n[2] Blinded data: {len(df)} rows, columns: {list(df.columns)}")
print(f"    Models: {sorted(df.model.unique())}")

# 3. No sensitive columns
sensitive = ["hallucination_detected", "flagged_terms", "target_token",
             "dangerous_reasoning_hallucination", "flagged_term_count"]
leaked = [c for c in sensitive if c in df.columns]
print(f"\n[3] Sensitive columns: {'None (OK)' if not leaked else leaked}")

# 4. No model name leaks
print("\n[4] Model name leak check:")
for col in df.columns:
    if df[col].dtype == "object":
        for kw in ["anthropic", "gemini", "openai", "openrouter", "claude", "llama", "gpt"]:
            m = df[col].str.contains(kw, case=False, na=False).sum()
            if m > 0:
                print(f"  LEAK: '{col}' has {m} matches for '{kw}'")
print("  Done.")

# 5. Stratification
df_full = pd.read_csv(OUTPUT_DIR / "rater_sample_100.csv")
ct = df_full.groupby(["model", "condition", "vignette_length"]).size()
print(f"\n[5] Cell sizes: min={ct.min()}, max={ct.max()}, all 4-5: {all(4 <= v <= 5 for v in ct.values)}")
print(f"    By model: {dict(df_full.groupby('model').size())}")
print(f"    By condition: {dict(df_full.groupby('condition').size())}")
print(f"    By length: {dict(df_full.groupby('vignette_length').size())}")

# 6. Excel files — same cases, different orders
print("\n[6] Excel files:")
case_orders = []
for i in range(1, 5):
    wb = load_workbook(OUTPUT_DIR / f"Psychiatrist_{i}_rating_sheet.xlsx", read_only=True)
    ws = wb["Rating Sheet"]
    ids = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
    case_orders.append(ids)
    print(f"  Psychiatrist_{i}: {len(ids)} cases, order: {ids[:5]}...")

sets = [set(o) for o in case_orders]
all_same = all(s == sets[0] for s in sets)
orders_different = len(set(tuple(o) for o in case_orders)) == 4
print(f"  All same cases: {all_same}")
print(f"  All different orders: {orders_different}")

print("\n" + "=" * 60)
print("ALL VALIDATIONS COMPLETE")
print("=" * 60)
