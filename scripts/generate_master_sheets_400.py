#!/usr/bin/env python3
"""
Generate Master Sheets for 400-Case Human-AI Rating Study
==========================================================

Creates 4 master sheets for tracking and analysis:
  1. 01_progress_tracker.xlsx — Track submission status
  2. 02_case_metadata_reference.xlsx — All 400 cases with anonymized info
  3. 03_rating_consolidation_template.xlsx — Combine ratings post-submission
  4. 04_analysis_sheet.xlsx — Full analysis with Kappa templates

Design:
  - 400 unique cases, each rated by 1 psychiatrist
  - 400 total ratings (4 psychiatrists × 100 cases)
  - Each psychiatrist gets 100 different cases
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
SOURCE_CSV = BASE / "04_results/human_validation/rater_sample_400_blinded.csv"
KEY_CSV = BASE / "04_results/human_validation/rater_sample_400_key.csv"
OUTPUT_DIR = BASE / "04_results/human_validation/master_sheets"

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="2F5496")
INFO_FONT = Font(size=10)
BODY_FONT = Font(size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INFO_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid")

# ── Helper Functions ────────────────────────────────────────────────────────

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    """Set cell value with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = THIN_BORDER

def write_progress_tracker():
    """Create progress tracker sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress Tracker"

    # Title
    set_cell(ws, 1, 1, "PROGRESS TRACKER — 400-CASE HUMAN-AI RATING", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, 2, 1, "", None, None, None, False)

    # Study info
    info_row = 4
    set_cell(ws, info_row, 1, "Study Design:", INFO_FONT)
    set_cell(ws, info_row, 2, "400 unique cases × 4 psychiatrists = 400 total ratings", BODY_FONT)
    set_cell(ws, info_row+1, 1, "Each psychiatrist rates 100 DIFFERENT cases", BODY_FONT)
    set_cell(ws, info_row+2, 1, "Cases are in DIFFERENT random order per psychiatrist", BODY_FONT)
    set_cell(ws, info_row+3, 1, "", None, None, None, False)

    # Rater info
    set_cell(ws, info_row+5, 1, "Raters:", INFO_FONT)
    rater_names = ["Psychiatrist_1", "Psychiatrist_2", "Psychiatrist_3", "Psychiatrist_4"]
    for i, rater in enumerate(rater_names, start=1):
        set_cell(ws, info_row+5+i, 1, f"  {rater}: 100 cases", BODY_FONT)

    set_cell(ws, info_row+11, 1, "", None, None, None, False)

    # Progress table
    set_cell(ws, info_row+13, 1, "PROGRESS", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+13, 2, "Rater", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+13, 3, "Cases Submitted", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+13, 4, "Status", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)

    for i, rater in enumerate(rater_names, start=1):
        set_cell(ws, info_row+14+i, 1, rater, BODY_FONT, INFO_FILL, LEFT_ALIGN)
        set_cell(ws, info_row+14+i, 2, "100", BODY_FONT, INFO_FILL, CENTER_ALIGN)
        set_cell(ws, info_row+14+i, 3, "Submitted", BODY_FONT, INFO_FILL, CENTER_ALIGN)
        set_cell(ws, info_row+14+i, 4, "✓", BODY_FONT, INFO_FILL, CENTER_ALIGN)

    # Summary
    set_cell(ws, info_row+20, 1, "SUMMARY", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+21, 1, "Total cases:", INFO_FONT)
    set_cell(ws, info_row+21, 2, "400", BODY_FONT)
    set_cell(ws, info_row+22, 1, "Total ratings:", INFO_FONT)
    set_cell(ws, info_row+22, 2, "400", BODY_FONT)
    set_cell(ws, info_row+23, 1, "Raters:", INFO_FONT)
    set_cell(ws, info_row+23, 2, "4", BODY_FONT)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 15
    for row in range(14, 19):
        ws.row_dimensions[row].height = 25

    wb.save(OUTPUT_DIR / "01_progress_tracker.xlsx")
    print("  ✅ 01_progress_tracker.xlsx")

def write_case_metadata_reference():
    """Create case metadata reference sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Metadata"

    # Load data
    df = pd.read_csv(SOURCE_CSV)
    df_key = pd.read_csv(KEY_CSV)

    # Create mapping
    model_map = dict(zip(df_key['anonymized_model'], df_key['actual_model']))
    df['model'] = df['model'].map(model_map)

    # Select columns
    columns = ['case_seq_id', 'case_id', 'model', 'condition', 'vignette_length', 'target_token']
    df = df[columns]

    # Write headers
    headers = ['Case Seq', 'Case ID', 'Model', 'Condition', 'Length', 'Fabricated Term']
    for col_idx, header in enumerate(headers, start=1):
        set_cell(ws, 1, col_idx, header, HEADER_FONT, HEADER_FILL, CENTER_ALIGN)

    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        set_cell(ws, row_idx, 1, row['case_seq_id'], BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 2, row['case_id'], BODY_FONT, ALT_ROW_FILL, LEFT_ALIGN)
        set_cell(ws, row_idx, 3, row['model'], BODY_FONT, ALT_ROW_FILL, LEFT_ALIGN)
        set_cell(ws, row_idx, 4, row['condition'], BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 5, row['vignette_length'], BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 6, row['target_token'], BODY_FONT, ALT_ROW_FILL, LEFT_ALIGN)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40

    wb.save(OUTPUT_DIR / "02_case_metadata_reference.xlsx")
    print("  ✅ 02_case_metadata_reference.xlsx")

def write_consolidation_template():
    """Create rating consolidation template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidation"

    # Title
    set_cell(ws, 1, 1, "RATING CONSOLIDATION TEMPLATE", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, 2, 1, "", None, None, None, False)

    # Instructions
    set_cell(ws, 4, 1, "Instructions:", INFO_FONT)
    set_cell(ws, 5, 1, "1. After all psychiatrists complete their ratings,", BODY_FONT)
    set_cell(ws, 6, 1, "   paste the Hallucination ratings into this template.", BODY_FONT)
    set_cell(ws, 7, 1, "2. Each row represents one case rated by one psychiatrist.", BODY_FONT)
    set_cell(ws, 8, 1, "3. Use the Case ID to match cases across psychiatrists.", BODY_FONT)
    set_cell(ws, 9, 1, "", None, None, None, False)

    # Headers
    headers = ['Case ID', 'Psychiatrist', 'Hallucination', 'Notes']
    for col_idx, header in enumerate(headers, start=1):
        set_cell(ws, 11, col_idx, header, HEADER_FONT, HEADER_FILL, CENTER_ALIGN)

    # Write placeholder rows
    for row_idx in range(12, 112):
        set_cell(ws, row_idx, 1, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 2, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 3, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 4, "", BODY_FONT, ALT_ROW_FILL, LEFT_ALIGN)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40

    wb.save(OUTPUT_DIR / "03_rating_consolidation_template.xlsx")
    print("  ✅ 03_rating_consolidation_template.xlsx")

def write_analysis_sheet():
    """Create analysis sheet with Kappa templates."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # Title
    set_cell(ws, 1, 1, "ANALYSIS SHEET — Kappa Calculation", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, 2, 1, "", None, None, None, False)

    # Study info
    info_row = 4
    set_cell(ws, info_row, 1, "Study Design:", INFO_FONT)
    set_cell(ws, info_row, 2, "400 unique cases × 4 psychiatrists = 400 total ratings", BODY_FONT)
    set_cell(ws, info_row+1, 1, "Each psychiatrist rates 100 DIFFERENT cases", BODY_FONT)
    set_cell(ws, info_row+2, 1, "", None, None, None, False)

    # Kappa calculation
    set_cell(ws, info_row+4, 1, "Kappa Calculation:", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+5, 1, "Use calculate_kappa_4raters.py to compute:", BODY_FONT)
    set_cell(ws, info_row+6, 1, "  - Cohen's Kappa for each pair of psychiatrists (6 pairs)", BODY_FONT)
    set_cell(ws, info_row+7, 1, "  - Fleiss' Kappa for 4-rater agreement", BODY_FONT)
    set_cell(ws, info_row+8, 1, "", None, None, None, False)

    # Input data
    set_cell(ws, info_row+10, 1, "Input Data:", INFO_FONT)
    set_cell(ws, info_row+11, 1, "Paste the Hallucination ratings from consolidation template here:", BODY_FONT)
    set_cell(ws, info_row+12, 1, "", None, None, None, False)

    # Data table
    set_cell(ws, info_row+14, 1, "Case ID", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+14, 2, "Psychiatrist_1", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+14, 3, "Psychiatrist_2", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+14, 4, "Psychiatrist_3", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    set_cell(ws, info_row+14, 5, "Psychiatrist_4", HEADER_FONT, HEADER_FILL, CENTER_ALIGN)

    # Placeholder rows
    for row_idx in range(15, 115):
        set_cell(ws, row_idx, 1, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 2, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 3, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 4, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)
        set_cell(ws, row_idx, 5, "", BODY_FONT, ALT_ROW_FILL, CENTER_ALIGN)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    wb.save(OUTPUT_DIR / "04_analysis_sheet.xlsx")
    print("  ✅ 04_analysis_sheet.xlsx")

def generate_all_sheets():
    """Generate all master sheets."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("GENERATING MASTER SHEETS")
    print("=" * 70)
    print()

    write_progress_tracker()
    write_case_metadata_reference()
    write_consolidation_template()
    write_analysis_sheet()

    print()
    print("=" * 70)
    print("ALL MASTER SHEETS GENERATED")
    print("=" * 70)
    print("Location:", OUTPUT_DIR)
    print()
    print("Files:")
    print("  01_progress_tracker.xlsx          — Track submission status")
    print("  02_case_metadata_reference.xlsx   — Case info (give to psychiatrists)")
    print("  03_rating_consolidation_template.xlsx — Combine ratings post-submission")
    print("  04_analysis_sheet.xlsx            — Full analysis with Kappa templates")
    print()
    print("Workflow:")
    print("  1. Give each psychiatrist their rating sheet + case metadata reference")
    print("  2. Track progress with progress tracker")
    print("  3. After all submitted, paste ratings into consolidation template")
    print("  4. Use analysis sheet for Kappa and agreement statistics")
    print("  5. Run calculate_kappa_4raters.py for automated Kappa calculation")
    print()

if __name__ == "__main__":
    generate_all_sheets()