#!/usr/bin/env python3
"""Final validation of all output files."""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

OUT = Path("04_results/human_validation")
MS = OUT / "master_sheets"

all_files = [
    OUT / "rater_sample_100_blinded.csv",
    OUT / "Psychiatrist_1_rating_sheet.xlsx",
    OUT / "Psychiatrist_2_rating_sheet.xlsx",
    OUT / "Psychiatrist_3_rating_sheet.xlsx",
    OUT / "Psychiatrist_4_rating_sheet.xlsx",
    MS / "01_progress_tracker.xlsx",
    MS / "02_case_metadata_reference.xlsx",
    MS / "03_rating_consolidation_template.xlsx",
    MS / "04_analysis_sheet.xlsx",
]

print("=== ALL FILES ===")
for f in all_files:
    size = f.stat().st_size if f.exists() else 0
    status = "OK" if f.exists() else "MISSING"
    print(f"  {status}: {f.name} ({size:,} bytes)")

df = pd.read_csv(OUT / "rater_sample_100_blinded.csv")
print(f"\n=== BLINDED CSV ===")
print(f"  Rows: {len(df)}, Cols: {list(df.columns)}")
print(f"  Has target_token: {'target_token' in df.columns}")

print(f"\n=== EXCEL RATING SHEETS ===")
for i in range(1, 5):
    wb = load_workbook(OUT / f"Psychiatrist_{i}_rating_sheet.xlsx", read_only=True)
    ws = wb["Rating Sheet"]
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    print(f"  Psychiatrist_{i}: {headers}")
    has_token = "Token" in headers
    print(f"    Has Token column: {has_token}")
    wb.close()

print(f"\n=== MASTER SHEETS ===")
for f in ["02_case_metadata_reference.xlsx", "03_rating_consolidation_template.xlsx"]:
    wb = load_workbook(MS / f, read_only=True)
    ws = wb.active
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
    print(f"  {f}: {headers[:8]}...")
    wb.close()

print("\n=== ALL VALIDATIONS COMPLETE ===")
